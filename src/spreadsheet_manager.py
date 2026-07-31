"""Criação, formatação e atualização automática da planilha Excel."""
import logging
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.config import (
    CURRENCY_PAIRS,
    MAX_HISTORY_ROWS,
    SHEET_DASHBOARD,
    SHEET_HISTORY,
    SPREADSHEET_PATH,
)
from src.data_fetcher import Quote

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E78")
UP_FILL = PatternFill("solid", fgColor="C6EFCE")
UP_FONT = Font(color="006100", bold=True)
DOWN_FILL = PatternFill("solid", fgColor="FFC7CE")
DOWN_FONT = Font(color="9C0006", bold=True)
NEUTRAL_FILL = PatternFill("solid", fgColor="FFEB9C")
NEUTRAL_FONT = Font(color="9C6500", bold=True)
THIN_BORDER = Border(*(Side(style="thin", color="B7B7B7"),) * 4)


def _style_header_row(ws: Worksheet, row: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _autofit_columns(ws: Worksheet, widths: dict[str, int]) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _ensure_workbook() -> Workbook:
    if SPREADSHEET_PATH.exists():
        return load_workbook(SPREADSHEET_PATH)

    wb = Workbook()
    dashboard = wb.active
    dashboard.title = SHEET_DASHBOARD
    wb.create_sheet(SHEET_HISTORY)
    logger.info("Nova planilha criada em %s", SPREADSHEET_PATH)
    return wb


def _write_dashboard(ws: Worksheet, quotes: dict[str, Quote]) -> None:
    ws.delete_rows(1, ws.max_row)

    ws.merge_cells("A1:D1")
    ws["A1"] = "Painel de Cotações — Atualização Automática"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws["A2"] = f"Última atualização: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A2"].font = Font(italic=True, size=9, color="666666")

    headers = ["Ativo", "Preço (R$)", "Variação (%)", "Situação"]
    header_row = 4
    for col, text in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=text)
    _style_header_row(ws, header_row, len(headers))

    row = header_row + 1
    for quote in quotes.values():
        ws.cell(row=row, column=1, value=quote["label"]).border = THIN_BORDER
        price_cell = ws.cell(row=row, column=2, value=round(quote["price"], 2))
        price_cell.number_format = '"R$" #,##0.00'
        price_cell.border = THIN_BORDER

        variation = quote["variation_pct"]
        variation_cell = ws.cell(row=row, column=3, value=round(variation, 2))
        variation_cell.number_format = "0.00\"%\""
        variation_cell.border = THIN_BORDER

        status_cell = ws.cell(row=row, column=4)
        status_cell.border = THIN_BORDER
        status_cell.alignment = Alignment(horizontal="center")
        if variation > 0:
            status_cell.value = "▲ Alta"
            for cell in (variation_cell, status_cell):
                cell.fill = UP_FILL
                cell.font = UP_FONT
        elif variation < 0:
            status_cell.value = "▼ Baixa"
            for cell in (variation_cell, status_cell):
                cell.fill = DOWN_FILL
                cell.font = DOWN_FONT
        else:
            status_cell.value = "► Estável"
            for cell in (variation_cell, status_cell):
                cell.fill = NEUTRAL_FILL
                cell.font = NEUTRAL_FONT

        row += 1

    _autofit_columns(ws, {"A": 28, "B": 16, "C": 16, "D": 14})
    ws.freeze_panes = "A5"


def _append_history(ws: Worksheet, quotes: dict[str, Quote]) -> None:
    is_new = ws.max_row <= 1 and ws.cell(row=1, column=1).value is None
    if is_new:
        headers = ["Data/Hora"] + [q["label"] for q in quotes.values()]
        for col, text in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=text)
        _style_header_row(ws, 1, len(headers))
        ws.freeze_panes = "A2"
        widths = {"A": 20}
        for i in range(len(quotes)):
            widths[get_column_letter(i + 2)] = 20
        _autofit_columns(ws, widths)

    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=1, value=datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    for col, quote in enumerate(quotes.values(), start=2):
        cell = ws.cell(row=new_row, column=col, value=round(quote["price"], 2))
        cell.number_format = '"R$" #,##0.00'
        cell.border = THIN_BORDER

    if new_row - 1 > MAX_HISTORY_ROWS:
        ws.delete_rows(2, new_row - 1 - MAX_HISTORY_ROWS)


def _rebuild_chart(ws_history: Worksheet, wb: Workbook) -> None:
    for chart in list(ws_history._charts):
        ws_history._charts.remove(chart)

    max_row = ws_history.max_row
    if max_row < 2:
        return

    chart = LineChart()
    chart.title = "Evolução das Cotações"
    chart.style = 12
    chart.y_axis.title = "Preço (R$)"
    chart.x_axis.title = "Data/Hora"
    chart.height = 10
    chart.width = 24

    num_series = len(CURRENCY_PAIRS)
    data = Reference(
        ws_history, min_col=2, max_col=1 + num_series, min_row=1, max_row=max_row
    )
    categories = Reference(ws_history, min_col=1, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    anchor_col = get_column_letter(3 + num_series)
    ws_history.add_chart(chart, f"{anchor_col}2")


def update_spreadsheet(quotes: dict[str, Quote]) -> None:
    """Atualiza a planilha em disco sem precisar abri-la no Excel."""
    if not quotes:
        logger.warning("Nenhuma cotação recebida, planilha não foi atualizada.")
        return

    wb = _ensure_workbook()
    dashboard = wb[SHEET_DASHBOARD]
    history = wb[SHEET_HISTORY]

    _write_dashboard(dashboard, quotes)
    _append_history(history, quotes)
    _rebuild_chart(history, wb)

    try:
        wb.save(SPREADSHEET_PATH)
        logger.info("Planilha atualizada com sucesso: %s", SPREADSHEET_PATH)
    except PermissionError:
        logger.error(
            "Não foi possível salvar: o arquivo %s está aberto em outro programa. "
            "Feche-o para que a próxima atualização automática seja aplicada.",
            SPREADSHEET_PATH,
        )
